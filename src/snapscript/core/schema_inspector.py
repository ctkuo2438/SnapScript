'''
conver the user's input csv or excel file into a SchemaReport for prompt_builder to use when generating prompts. 
This module is focused on schema inspection and reporting, not on data transformation or cleaning.

Streamlit UI / CLI
  -> schema_inspector.inspect(...)
      -> SchemaReport
      
Ex: 
SchemaReport(
    filename="orders.csv",
    file_type="csv",
    row_count=82,
    file_size_bytes=12345,
    columns=[
        ColumnInfo(
            name="quantity",
            dtype="object",
            null_count=3,
            unique_count=8,
            sample_values=["1", "2", "10", "two", "-1"],
        ),
        ColumnInfo(
            name="unit_price",
            dtype="object",
            null_count=2,
            unique_count=50,
            sample_values=["$216.27", "USD 55.20", "121.25", "N/A"],
        ),
    ],
    sample_rows=[
        {"order_id": "ORD-1001", "quantity": "1", "unit_price": "$216.27"},
        ...
    ],
    encoding="utf-8",
    sheet_names=[],
)
'''

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

import chardet # for encoding detection of csv files
import pandas as pd
from openpyxl import load_workbook

from snapscript.config import AppConfig
from snapscript.core.models import ColumnInfo, SchemaReport

# create a SnapScript-specific error hierarchy for schema inspection issues
# inherit from python build-in Exception class
class SchemaInspectionError(Exception):
    """Base error for schema inspection failures."""


class MissingInputFileError(SchemaInspectionError):
    """Raised when the requested input file does not exist."""


class UnsupportedFileTypeError(SchemaInspectionError):
    """Raised when the input file extension is not supported."""


class UnreadableFileError(SchemaInspectionError):
    """Raised when a supported file cannot be read."""


class InputFileTooLargeError(SchemaInspectionError):
    """Raised when the input file exceeds configured size limits."""


SUPPORTED_EXTENSIONS = frozenset({".csv", ".xlsx", ".xls"})


# main entry point for schema inspection, called by the API handler after file upload, 
#   returns a SchemaReport or raises a SchemaInspectionError
def inspect(path: Path, sheet: str | None = None) -> SchemaReport:
    config = AppConfig()
    input_path = Path(path)
    # 1. validate file existence, type, and size before attempting to read
    _validate_path(input_path, config)

    suffix = input_path.suffix.lower()
    try:
        # 2. identify the file extension is csv or excel
        if suffix == ".csv":
            # 3. read the csv file sample and metadata to build the SchemaReport
            return _inspect_csv(input_path, config)
        # 4. if excel, read the specified sheet (or first sheet by default) sample and metadata to build the SchemaReport
        return _inspect_excel(input_path, sheet, config)
    
    # if is schema inspection error, then raise it directly
    except SchemaInspectionError:
        raise
    # otherwise, wrap any unexpected exceptions as UnreadableFileError, ex: pandas/openpyxl/chardet related errors
    except Exception as exc:
        raise UnreadableFileError(f"Could not read file: {input_path}") from exc

# TODO: add Multi-file schema: inspect_many(inputs: list[InputFileSpec]) -> MultiFileSchemaReport

# helper functions for schema inspection, focused on reading file metadata and sample data without loading the entire dataset into memory, 
#   to avoid performance issues with large files.
def _validate_path(path: Path, config: AppConfig) -> None:
    if not path.exists():
        raise MissingInputFileError(f"File not found: {path}")
    if not path.is_file():
        raise UnreadableFileError(f"Input path is not a file: {path}")
    if path.suffix.lower() not in SUPPORTED_EXTENSIONS: # if not .csv/.xlsx/.xls, then raise unsupported file type error
        raise UnsupportedFileTypeError(f"Unsupported file type: {path.suffix}")

    file_size = path.stat().st_size # get the file size in bytes
    if file_size > config.max_input_file_size_bytes: # exceed 500 mb, then raise input file too large error
        raise InputFileTooLargeError(
            f"File too large: {file_size} bytes exceeds "
            f"{config.max_input_file_size_bytes} bytes"
        )

'''
generate a schema report that contains metadata about the input file: 
filename, file_type, row_count, file_size_bytes, columns, sample_rows, encoding, sheet_names (for excel).
'''
def _inspect_csv(path: Path, config: AppConfig) -> SchemaReport:
    encoding = _detect_encoding(path) # may not be utf-8, need to detect encoding first to avoid read errors
    sample = pd.read_csv(path, nrows=config.schema_inspect_rows, encoding=encoding) # read only the 1000 rows
    sample = _truncate_columns(sample, config)

    return SchemaReport(
        filename=path.name,
        file_type="csv",
        row_count=_count_csv_rows(path, encoding),
        file_size_bytes=path.stat().st_size,
        columns=_build_columns(sample),
        sample_rows=_sample_rows(sample, config),
        encoding=encoding,
        sheet_names=[],
    )


def _inspect_excel(path: Path, sheet: str | None, config: AppConfig) -> SchemaReport:
    excel_file = pd.ExcelFile(path)
    sheet_names = [str(name) for name in excel_file.sheet_names]
    
    # if no sheet is specified, use the first sheet by default
    #   Otherwise, validate that the requested sheet exists
    selected_sheet = sheet or sheet_names[0]
    if selected_sheet not in sheet_names:
        raise UnreadableFileError(f"Sheet not found: {selected_sheet}")

    sample = pd.read_excel(
        excel_file,
        sheet_name=selected_sheet,
        nrows=config.schema_inspect_rows,
    )
    sample = _truncate_columns(sample, config)

    return SchemaReport(
        filename=path.name,
        file_type=path.suffix.lower().lstrip("."), # xlsx or xls
        row_count=_count_excel_rows(path, selected_sheet, sample),
        file_size_bytes=path.stat().st_size,
        columns=_build_columns(sample),
        sample_rows=_sample_rows(sample, config),
        encoding="utf-8",
        sheet_names=sheet_names,
    )


def _detect_encoding(path: Path) -> str:
    with path.open("rb") as file:
        raw = file.read(64 * 1024) # read the first 64kb for encoding detection

    detected = chardet.detect(raw) # detect the encoiding using chardet
    encoding = detected.get("encoding")
    return encoding or "utf-8" # if not detected, default to utf-8


# count the number of rows in the csv file without loading the entire file into memory, 
#   by iterating through the file with csv.reader and counting lines
def _count_csv_rows(path: Path, encoding: str) -> int:
    with path.open("r", encoding=encoding, newline="") as file:
        reader = csv.reader(file)
        try:
            next(reader) # skip header row
        except StopIteration:
            return 0
        return sum(1 for _ in reader)


# count the number of rows in the excel sheet without loading the entire sheet into memory
def _count_excel_rows(path: Path, sheet: str, sample: pd.DataFrame) -> int:
    if path.suffix.lower() != ".xlsx":
        return len(sample) # for .xls files, directly return the number of rows

    # if .xlsx file, use openpyxl to read the sheet in read-only mode and get the max_row property, 
    #   which is more efficient than loading with pandas and can handle larger files without memory issues
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet]
        return max(worksheet.max_row - 1, 0) # -1 to exclude header row
    finally:
        workbook.close()


def _truncate_columns(df: pd.DataFrame, config: AppConfig) -> pd.DataFrame:
    truncated = df.copy()
    # convert the column names to string and truncate to max_column_name_chars to avoid issues with very long column names in the prompt
    truncated.columns = [
        str(column)[: config.max_column_name_chars] for column in truncated.columns
    ]
    return truncated


'''
create ColumnInfo objects for each column in the dataframe, including name, dtype, null count, unique count, and sample values.
these info will be used in the prompt to help the LLM understand the structure and content of the data when generating code

Ex:
ColumnInfo(
        name="order_id",
        dtype="object",
        null_count=0,
        unique_count=82,
        sample_values=["ORD-1001", "ORD-1002", "ORD-1003", "ORD-1004", "ORD-1005"],
    ),
'''
def _build_columns(df: pd.DataFrame) -> list[ColumnInfo]:
    columns: list[ColumnInfo] = []
    for column_name in df.columns:
        series = df[column_name] # get the column as a pandas Series
        non_null = series.dropna() # drop null values for unique count and sample values to avoid counting NaN as a unique value and including it in samples
        columns.append(
            ColumnInfo(
                name=str(column_name),
                dtype=str(series.dtype),
                null_count=int(series.isna().sum()), # count the number of null values in the column
                unique_count=int(series.nunique(dropna=True)), # count the number of unique values in the column, excluding nulls
                sample_values=_sample_values(non_null),
            )
        )
    return columns


# get up to 5 unique sample values from the column to include in the prompt, 
#   which can help the LLM understand the content of the column when generating code
def _sample_values(series: pd.Series[Any]) -> list[str]:
    values: list[str] = []
    for value in series.drop_duplicates().head(5).tolist():
        values.append(str(value))
    return values


# get up to schema_sample_rows rows of sample data as a list of dicts, 
#   to include in the prompt for LLM to understand the data when generating code
def _sample_rows(df: pd.DataFrame, config: AppConfig) -> list[dict[str, object]]:
    records = df.head(config.schema_sample_rows).where(pd.notna(df), None).to_dict(
        orient="records"
    )
    return [dict(record) for record in records]
